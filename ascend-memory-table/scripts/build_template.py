# -*- coding: utf-8 -*-
"""
生成 ascend-memory-table 的标准模板 template.xlsx。

模板中：
- A 列：行标签（固定）
- B 列：留空（由 fill_table.py 按模型/部署参数填入）
- C 列：计算公式 + 源码锚点（已核对固定，fill_table.py 仅在含动态数值的行追加具体数字）

公式口径见 references/formulas.md（对齐本地 vllm d6dbdb9b0 / vllm-ascend f5b5514af）。
重新生成模板：python build_template.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from pathlib import Path

HEADER = PatternFill("solid", fgColor="D9EAF7")
YELLOW = PatternFill("solid", fgColor="FFF7D6")
COL_HDR = PatternFill("solid", fgColor="E8F1FF")

# (row, colA, colC)  colB 留空由脚本填；colC 为已核对的公式/源码说明
ROWS = [
    (1,  "自定义配置项", None),
    (2,  "B 平均请求长度(输入+输出）", "用户输入（输入+输出 tokens），决定「最大并发（KV容量理论）」行"),
    (3,  "max-model-len", "--max-model-len [vllm/config/model.py]"),
    (4,  "gpu-memory-utilization", "--gpu-memory-utilization [CacheConfig.gpu_memory_utilization, vllm/config/cache.py]"),
    (5,  "", None),
    (6,  "模型", "用户输入 / HF model id / 本地模型目录名"),
    (7,  "L 模型层数", "config.num_hidden_layers [HF config.json]"),
    (8,  "L_sparse 稀疏/index层数", "sum(sparse_attention_freq==1)；无 sparse → 0 [MiniMaxM3TextConfig.sparse_attention_config]"),
    (9,  "Hkv kv头数", "全局 num_key_value_heads；TP 后 Hkv_local=max(1,Hkv//TP)；MLA(DeepSeek) 不使用 Hkv（用 kv_lora_rank）[HF config / vLLM TP]"),
    (10, "D head_dim", "config.head_dim；DeepSeek MLA 无此字段（用 qk_nope+qk_rope / v_head_dim）[HF config]"),
    (11, "IdxD index_head_dim", "sparse_attention_config.sparse_index_dim；indexer key-only；无 → 0 [HF config]"),
    (12, "bytes_per_elem", "KV/Index dtype 字节；BF16=2，C8/INT8=1 [CacheConfig.cache_dtype → get_dtype_size]"),
    (13, "hidden_size", "config.hidden_size [HF config]"),
    (14, "num_experts_per_tok", "config.num_experts_per_tok (top-k) [HF config]"),
    (15, "num_experts", "config.num_local_experts [HF config]"),
    (16, "experts_per_device", "num_experts // EP [推导]"),
    (17, "", None),
    (18, "部署参数", None),
    (19, "TP", "--tensor-parallel-size [vllm/config/parallel.py]"),
    (20, "DP", "--data-parallel-size [vllm/config/parallel.py]"),
    (21, "EP", "EP=TP×DP（开 --enable-expert-parallel）[vllm/distributed; vLLM-Ascend MoE EP]"),
    (22, "", None),
    (23, "实测显存", "对齐 worker.py warmup 日志；蓝底=实测，绿底=理论估算"),
    (24, "权重 (GiB)", "model_runner.model_memory_usage（实测）；理论 W8A8 见行39 [worker.py::load_model]"),
    (25, "激活占用显存汇总 (GiB)", "torch_peak_increase（profile_run 实测）；理论≈1.25×(2·T·H+T·topk·H+T·n_h_local·D+T·H+2·(T·topk/EP)·H)×2B [worker.py:563; mem_utils.py:310]"),
    (26, "激活占用显存汇总 (MB)", "act_bytes / 1e6（十进制 MB）[推导，对齐部分日志]"),
    (27, "non-torch memory (GiB)", "non_torch_increase = after_profile.non_torch − before_create.non_torch（HCCL/CANN/驱动）[mem_utils.py:311; worker.py:569]"),
    (28, "NPU graph memory (GiB)", "npugraph_memory_bytes = capture_model()（cudagraph 捕获，不进 Current KV）[worker.py:710,727]"),
    (29, "Current KV cache memory (GiB)", "available_kv = requested − (W + peak_act + non_torch) ※不含 graph [worker.py:581]"),
    (30, "kv-cache-memory fit requested (GiB)", "requested − (W + peak_act + non_torch + graph) − 150MiB [worker.py:720,728]"),
    (31, "kv-cache-memory fully utilize free (GiB)", "init_free − (W + peak_act + non_torch + graph) − 150MiB [worker.py:729]"),
    (32, "num_blocks", "available_kv // page // layers（hybrid 时 // max(L,L_sparse)）；page=2·block·Hkv_local·D·dtype [kv_cache_utils.py:1009 get_num_blocks; :1380]"),
    (33, "block_size", "cache_config.block_size（Ascend refresh_block_size 默认 128）[vllm_ascend/utils.py:1241]"),
    (34, "GPU KV cache size (tokens)", "max_concurrency × max_model_len [kv_cache_utils.py:1833 get_kv_cache_capacity]"),
    (35, "Maximum concurrency for max_model_len", "num_blocks / blocks_per_req；blocks_per_req=Σ cdiv(max_model_len, group.block_size) [kv_cache_utils.py:951,958]"),
    (36, "", None),
    (37, "显存占用与并发", None),
    (38, "Tensor", "计算公式（vLLM / vLLM-Ascend 源码口径）"),
    (39, "权重 (GiB)", "W8A8 = K_emb_lm/TP·2 + K_norms·2 + K_gate·4 + (Q_attn[+Q_dense+Q_shared+Q_idx])/TP·1 + Q_experts/EP·1 [worker.py + 推导]"),
    (40, "激活占用显存汇总 (GiB)", "同行25 [worker.py]"),
    (41, "可用KV Cache (GiB)", "requested − W − act − non_torch（= 行29）[worker.py:581]"),
    (42, "单个token占用的kv cache(MiB)", "GQA: L × 2 × Hkv_local × D × dtype；MLA(DeepSeek): L × (kv_lora_rank+qk_rope_head_dim) × dtype（无 ×2）[AttentionSpec.real_page_size_bytes / block_size]"),
    (43, "单个token占用的index cache(MiB)", "L_sparse × 1 × IdxD × dtype（key-only，无 ×2）[MLAAttentionSpec; indexer.py:161]"),
    (44, "单个token占用的draft kv cache(MiB)", "MTP/NextN/EAGLE draft 层 KV = draft_layers × per_layer_main_kv；无 draft → 0 [vllm/v1/spec_decode]"),
    (45, "单个token合计cache(MiB)", "main + index + draft（每卡口径，已含 TP 分片）[推导]"),
    (46, "KV Cache 可容纳token数（日志块口径）", "max_concurrency × max_model_len（= 行34）[kv_cache_utils.py:1833]"),
    (47, "最大并发（KV容量理论）", "(available_kv / sum_bytes_per_token) / B [kv_cache_utils.py:958]"),
]

# B 列填黄色（用户输入/部署参数），表头填 HEADER，列头填 COL_HDR
YELLOW_ROWS = {2, 3, 4, 19, 20, 21}
HEADER_ROWS = {1, 6, 18, 23, 37}
COL_HDR_ROW = 38


def build(out_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    align = Alignment(wrap_text=True, vertical="center")
    for r, a, c in ROWS:
        ca = ws.cell(r, 1, a)
        cb = ws.cell(r, 2, None)
        cc = ws.cell(r, 3, c)
        for cell in (ca, cb, cc):
            cell.alignment = align
        if r in HEADER_ROWS:
            for col in (1, 2, 3):
                ws.cell(r, col).fill = HEADER
        if r == COL_HDR_ROW:
            for col in (1, 2, 3):
                ws.cell(r, col).fill = COL_HDR
        if r in YELLOW_ROWS:
            ws.cell(r, 2).fill = YELLOW
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 95
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    out = Path(__file__).parent / "template.xlsx"
    build(out)
    print(f"OK -> {out}")
