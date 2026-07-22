# -*- coding: utf-8 -*-
"""
Ascend NPU 显存填表脚本：输入 config.json + 拉起命令 → 输出 xlsx。

用法:
    python fill_table.py --config /path/to/config.json --serve "vllm serve ... --tensor-parallel-size 8 ..." \
        --hbm 64 --B 8192 --out filled.xlsx
    python fill_table.py --config-dir /path/to/model_dir --serve @serve.sh --hbm 64 --out filled.xlsx
    python fill_table.py --config config.json --serve "..." --warmup-log vllm.log --out filled.xlsx

公式口径见 references/formulas.md（对齐本地 vllm/vllm-ascend 源码）。
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shlex
import sys
import tempfile
from pathlib import Path
from urllib.error import URLError
from urllib.request import Request, urlopen

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill

from vllm_ascend_memory_formulas import (
    GiB, MiB, REDUNDANCY_BUFFER, KVLayout, MemoryBudget,
    estimate_peak_activation_bytes, local_num_kv_heads,
)

HEADER = PatternFill("solid", fgColor="D9EAF7")
YELLOW = PatternFill("solid", fgColor="FFF7D6")
COL_HDR = PatternFill("solid", fgColor="E8F1FF")
EST = PatternFill("solid", fgColor="E8F8E8")
WARN = PatternFill("solid", fgColor="FCE8E6")
MEAS = PatternFill("solid", fgColor="D6EAF8")  # measured (from log)


# ---------- HuggingFace / ModelScope auto-download ----------
# Resolves a model id (e.g. "MiniMaxAI/MiniMax-M3") or URL to a local file.
# Downloads ONLY small metadata files (config.json, quant_model_description.json), never weights.
# Sources (auto-fallback HF → ModelScope when HF unreachable):
#   - HuggingFace:  https://huggingface.co/<id>/resolve/<rev>/<file>
#   - ModelScope:   https://modelscope.cn/api/v1/models/<id>/repo?Revision=<rev>&FilePath=<file>
# Uses plain urllib (no huggingface_hub / modelscope SDK) so the skill stays self-contained.
HF_ENDPOINT = os.environ.get("HF_ENDPOINT", "https://huggingface.co")
HF_TOKEN = os.environ.get("HF_TOKEN") or os.environ.get("HUGGING_FACE_HUB_TOKEN")
MS_ENDPOINT = os.environ.get("MODELSCOPE_ENDPOINT", "https://modelscope.cn")
MS_TOKEN = os.environ.get("MODELSCOPE_API_TOKEN")


def _hf_resolve_id(s: str) -> str:
    """Accept 'owner/name', 'owner/name/revision', or a full HF URL; return 'owner/name' (+ optional @revision)."""
    s = s.strip()
    if s.startswith(("http://", "https://")):
        m = re.match(r"https?://[^/]+/([^/]+/[^/?#]+)(?:/(?:resolve|blob|tree)/([^/?#]+))?", s)
        if m:
            return m.group(1) + (f"@{m.group(2)}" if m.group(2) else "")
    return s


def _http_get(url: str, headers: dict, timeout: int = 30) -> bytes:
    req = Request(url, headers=headers)
    with urlopen(req, timeout=timeout) as r:
        return r.read()


def _save_json(data: bytes, out: Path) -> None:
    try:
        json.loads(data)  # validate
    except json.JSONDecodeError:
        raise ValueError("downloaded content is not valid JSON")
    out.write_bytes(data)


def _hf_fetch(model_id: str, filename: str, dest_dir: Path) -> Path:
    """Download a file from HuggingFace into dest_dir; return local path. Raises on failure."""
    dest_dir.mkdir(parents=True, exist_ok=True)
    safe = model_id.replace("/", "_").replace("@", "_rev_")
    out = dest_dir / f"{safe}__hf__{filename.replace('/', '_')}"
    if out.exists() and out.stat().st_size > 0:
        return out
    mid, _, rev = model_id.partition("@")
    rev = rev or "main"
    url = f"{HF_ENDPOINT}/{mid}/resolve/{rev}/{filename}"
    headers = {"User-Agent": "ascend-memory-table/1.0"}
    if HF_TOKEN:
        headers["Authorization"] = f"Bearer {HF_TOKEN}"
    data = _http_get(url, headers)
    _save_json(data, out)
    return out


def _ms_fetch(model_id: str, filename: str, dest_dir: Path) -> Path:
    """Download a file from ModelScope into dest_dir; return local path. Raises on failure."""
    dest_dir.mkdir(parents=True, exist_ok=True)
    safe = model_id.replace("/", "_").replace("@", "_rev_")
    out = dest_dir / f"{safe}__ms__{filename.replace('/', '_')}"
    if out.exists() and out.stat().st_size > 0:
        return out
    mid, _, rev = model_id.partition("@")
    rev = rev or "master"
    url = f"{MS_ENDPOINT}/api/v1/models/{mid}/repo?Revision={rev}&FilePath={filename}"
    headers = {"User-Agent": "ascend-memory-table/1.0"}
    if MS_TOKEN:
        headers["Authorization"] = f"Bearer {MS_TOKEN}"
    data = _http_get(url, headers)
    _save_json(data, out)
    return out


def download_file(model_id: str, filename: str, dest_dir: Path, source: str = "auto",
                  prefer: str | None = None) -> tuple[Path | None, str | None]:
    """
    Download a metadata file for a model id. Returns (local_path_or_None, source_used_or_None).
    source: 'hf' | 'ms' | 'auto'. A 404/missing file is not fatal → returns (None, None).
    `prefer` pins a source (used to reuse the source that worked for config.json).
    """
    mid = _hf_resolve_id(model_id)
    order = []
    if prefer:
        order.append(prefer)
    if source in ("auto", "hf") and "hf" not in order:
        order.append("hf")
    if source in ("auto", "ms") and "ms" not in order:
        order.append("ms")
    last_err = None
    for src in order:
        try:
            if src == "hf":
                return _hf_fetch(mid, filename, dest_dir), "hf"
            else:
                return _ms_fetch(mid, filename, dest_dir), "ms"
        except Exception as e:
            last_err = e
            # 404 / missing file → try next source; other errors also non-fatal for optional files
            continue
    return None, None


def download_config(model_id: str, dest_dir: Path, source: str = "auto") -> tuple[Path, str]:
    """Download config.json (required). Exits with helpful error if all sources fail."""
    path, used = download_file(model_id, "config.json", dest_dir, source=source)
    if path is None:
        sys.exit(
            f"[fill_table] 下载 config.json 失败（已尝试 {source}）: model_id={_hf_resolve_id(model_id)}\n"
            f"  排查: 1) 模型 id 是否正确 2) 私有/gated 仓库需设 HF_TOKEN / MODELSCOPE_API_TOKEN "
            f"3) 国内可设 HF_ENDPOINT=https://hf-mirror.com 或用 --source ms 4) 网络代理"
        )
    return path, used


# ---------- KV cache dtype (bytes_per_elem) auto-detection ----------
# Official rules (verified against upstream source):
#   1. vLLM default cache_dtype = "auto" → uses model dtype (BF16 → 2 bytes)
#      [vllm/config/cache.py: CacheConfig.cache_dtype = "auto"]
#   2. --kv-cache-dtype CLI flag overrides: fp8/fp8_e4m3/fp8_e5m2/int8 → 1 byte; else 2
#   3. vLLM-Ascend C8 (INT8 KV cache): driven by the checkpoint's quant_model_description.json
#      with "kv_cache_type": "C8" → overrides kv_cache_torch_dtype = torch.int8 (1 byte).
#      NOT enabled by --quantization ascend alone: W8A8 (weight only) keeps KV BF16;
#      W8A8C8 (weight + KV) sets kv_cache_type=C8. Activates automatically, no CLI flag needed.
#      [vllm_ascend/quantization/modelslim_config.py: kv_cache_type == "C8"]
def detect_bpe(deploy: dict, model_id: str | None, dest_dir: Path, source: str,
              prefer_source: str | None, local_config_dir: Path | None) -> tuple[int, str]:
    """
    Determine KV cache bytes-per-element. Returns (bpe, explanation).
    Priority: explicit --kv-cache-dtype > quant_model_description.json (C8) > vLLM default (BF16).
    """
    # 1) explicit --kv-cache-dtype in serve command
    kcd = deploy.get("kv_cache_dtype")
    if kcd:
        low = str(kcd).lower()
        if "8" in low:  # fp8, fp8_e4m3, fp8_e5m2, int8
            return 1, f"--kv-cache-dtype={kcd} → 1 byte/elem [vllm CacheConfig.cache_dtype]"
        return 2, f"--kv-cache-dtype={kcd} → 2 bytes/elem (BF16)"

    # 2) quant_model_description.json (Ascend C8 / INT8 KV cache)
    qd = _load_quant_desc(model_id, dest_dir, source, prefer_source, local_config_dir)
    if qd:
        kt = (qd.get("kv_cache_type") or qd.get("kv_quant_type") or "")
        kt_u = str(kt).upper()
        if "C8" in kt_u or "INT8" in kt_u:
            return 1, (f"quant_model_description.json: kv_cache_type={kt!r} → INT8 KV cache, "
                      f"1 byte/elem [vllm_ascend/quantization/modelslim_config.py]")
        return 2, (f"quant_model_description.json: kv_cache_type={kt!r} → BF16 KV cache, "
                  f"2 bytes/elem (W8A8 weight-only, no C8)")

    # 3) vLLM default
    return 2, "vLLM default cache_dtype='auto' → model dtype (BF16, 2 bytes/elem) [vllm/config/cache.py]"


def _load_quant_desc(model_id: str | None, dest_dir: Path, source: str,
                     prefer_source: str | None, local_config_dir: Path | None) -> dict | None:
    """Load quant_model_description.json (Ascend ModelSlim quant config) if present."""
    # local: sit next to config.json
    if local_config_dir is not None:
        local = local_config_dir / "quant_model_description.json"
        if local.exists():
            try:
                return json.loads(local.read_text(encoding="utf-8"))
            except Exception:
                return None
    # remote: download from the same source as config.json
    if model_id:
        path, _ = download_file(model_id, "quant_model_description.json", dest_dir,
                                source=source, prefer=prefer_source)
        if path is not None:
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                return None
    return None


# ---------- serve command parsing ----------
def parse_serve(cmd: str) -> dict:
    """Extract deploy params from a vllm serve / api_server command string."""
    d = {
        "TP": 1, "DP": 1, "EP": None, "util": 0.9, "max_model_len": None,
        "max_num_seqs": None, "max_num_batched_tokens": None,
        "block_size": 128, "quantization": None, "kv_cache_dtype": None,
        "enable_expert_parallel": False, "enable_async_scheduling": False,
    }
    tokens = shlex.split(cmd)
    def val(flag, cast=str):
        for i, t in enumerate(tokens):
            if t == flag and i + 1 < len(tokens):
                return cast(tokens[i + 1])
            if t.startswith(f"{flag}="):
                return cast(t.split("=", 1)[1])
        return None
    if v := val("--tensor-parallel-size", int): d["TP"] = v
    if v := val("--data-parallel-size", int): d["DP"] = v
    if v := val("--gpu-memory-utilization", float): d["util"] = v
    if v := val("--max-model-len", int): d["max_model_len"] = v
    if v := val("--max-num-seqs", int): d["max_num_seqs"] = v
    if v := val("--max-num-batched-tokens", int): d["max_num_batched_tokens"] = v
    if v := val("--block-size", int): d["block_size"] = v
    if v := val("--quantization", str): d["quantization"] = v
    if v := val("--kv-cache-dtype", str): d["kv_cache_dtype"] = v
    if "--enable-expert-parallel" in tokens: d["enable_expert_parallel"] = True
    if "--async-scheduling" in tokens: d["enable_async_scheduling"] = True
    d["EP"] = d["TP"] * d["DP"] if d["enable_expert_parallel"] else d["TP"]
    return d


# ---------- config.json parsing ----------
def load_config(path: Path) -> dict:
    cfg = json.loads(path.read_text(encoding="utf-8"))
    # unwrap text_config for VL models
    if "text_config" in cfg and isinstance(cfg["text_config"], dict):
        tc = cfg["text_config"]
    else:
        tc = cfg
    return {"raw": cfg, "text": tc, "vision": cfg.get("vision_config")}


def arch_from_config(cfg: dict) -> dict:
    """
    Family-aware architecture extraction. Normalizes field-name differences across
    HuggingFace/ModelScope config.json formats so the rest of the skill is family-agnostic.

    Supported families (verified against upstream config.json):
      - MiniMax-M3 (minimax_m3_vl): sparse_attention_config, num_local_experts,
        dense_intermediate_size, shared_intermediate_size, moe_layer_freq[list], MTP
      - Qwen3-MoE (qwen3_moe): num_experts, moe_intermediate_size, decoder_sparse_step,
        mlp_only_layers, no shared experts, no sparse index
      - DeepSeek-V2/V3 (deepseek_v2/v3): MLA (kv_lora_rank/qk_rope_head_dim),
        n_routed_experts, n_shared_experts, moe_intermediate_size, first_k_dense_replace
      - Dense Llama/Qwen2/GLM/etc.: no MoE, standard GQA
    """
    tc = cfg["text"]
    model_type = tc.get("model_type", "") or cfg.get("raw", {}).get("model_type", "") or ""
    L = tc["num_hidden_layers"]
    n_heads = tc["num_attention_heads"]
    Hkv = tc.get("num_key_value_heads", n_heads)  # MHA fallback
    H = tc["hidden_size"]
    vocab = tc["vocab_size"]
    tie = tc.get("tie_word_embeddings", False)

    # ---- head_dim: present for Qwen3/MiniMax; derived for DeepSeek MLA ----
    D = tc.get("head_dim")
    if D is None:
        # DeepSeek MLA: no head_dim; Q head = qk_nope+qk_rope, V head = v_head_dim.
        # For GQA-style per-token KV we don't use D under MLA (uses kv_lora_rank),
        # but keep a representative value for reporting / non-MLA fallback.
        D = (tc.get("qk_nope_head_dim", 0) + tc.get("qk_rope_head_dim", 0)) or tc.get("v_head_dim", 0) or 128

    # ---- MLA detection (DeepSeek-V2/V3) ----
    is_mla = "kv_lora_rank" in tc and "q_lora_rank" in tc
    kv_lora_rank = tc.get("kv_lora_rank", 0) or 0
    qk_rope_head_dim = tc.get("qk_rope_head_dim", 0) or 0
    qk_nope_head_dim = tc.get("qk_nope_head_dim", 0) or 0
    v_head_dim = tc.get("v_head_dim", 0) or 0
    q_lora_rank = tc.get("q_lora_rank", 0) or 0

    # ---- MoE experts: num_experts (Qwen3) / num_local_experts (MiniMax) / n_routed_experts (DeepSeek) ----
    num_experts = (tc.get("num_experts") or tc.get("num_local_experts")
                  or tc.get("n_routed_experts") or 0)
    num_experts_per_tok = tc.get("num_experts_per_tok", 0) or 0

    # ---- per-expert intermediate: moe_intermediate_size (Qwen3/DeepSeek) / intermediate_size (MiniMax routed) ----
    moe_intermediate = tc.get("moe_intermediate_size", 0) or 0
    routed_intermediate = moe_intermediate or tc.get("intermediate_size", 0) or 0

    # ---- dense (non-MoE) layer intermediate ----
    dense_intermediate = (tc.get("dense_intermediate_size", 0)  # MiniMax
                         or (tc.get("intermediate_size", 0) if not moe_intermediate else 0)  # dense-only models
                         or 0)

    # ---- shared experts (DeepSeek/MiniMax) ----
    n_shared = tc.get("n_shared_experts", 0) or 0
    shared_intermediate = (tc.get("shared_intermediate_size", 0)  # MiniMax
                           or moe_intermediate  # DeepSeek shared uses moe_intermediate_size
                           or 0)

    # ---- n_dense (first dense, rest MoE) ----
    if "first_k_dense_replace" in tc:                      # DeepSeek
        n_dense = int(tc["first_k_dense_replace"])
    elif isinstance(tc.get("moe_layer_freq"), list):       # MiniMax: list of 0/1
        freq = tc["moe_layer_freq"]
        n_dense = sum(1 for x in freq if x == 0) if freq else 0
    elif isinstance(tc.get("moe_layer_freq"), int):       # DeepSeek: every N-th layer MoE
        n_dense = 0 if tc["moe_layer_freq"] == 1 else (L // max(1, tc["moe_layer_freq"]))
    elif isinstance(tc.get("mlp_only_layers"), list) and tc.get("mlp_only_layers"):
        n_dense = len(tc["mlp_only_layers"])               # Qwen3: explicit dense-only list
    elif "decoder_sparse_step" in tc:
        n_dense = 0 if int(tc["decoder_sparse_step"]) == 1 else 0  # Qwen3: step=1 → all MoE
    else:
        n_dense = L if num_experts == 0 else 0             # dense model: all dense
    n_moe = L - n_dense

    # ---- sparse / index attention (MiniMax-M3) ----
    sac = tc.get("sparse_attention_config", {}) or {}
    sparse_freq = sac.get("sparse_attention_freq", [])
    L_sparse = sum(1 for x in sparse_freq if x == 1) if sparse_freq else 0
    IdxD = sac.get("sparse_index_dim", 0) or 0
    idx_heads = sac.get("sparse_num_index_heads", 0) or 0

    # ---- MTP / draft layers (MiniMax-M3 num_mtp_modules; DeepSeek-V3 num_nextn_predict_layers) ----
    draft_layers = (tc.get("num_mtp_modules", 0) or tc.get("num_nextn_predict_layers", 0) or 0)

    return {
        "L": L, "L_sparse": L_sparse,
        "n_heads": n_heads, "Hkv": Hkv, "D": D, "hidden_size": H,
        "vocab_size": vocab, "tie": tie,
        "is_mla": is_mla,
        "kv_lora_rank": kv_lora_rank, "qk_rope_head_dim": qk_rope_head_dim,
        "qk_nope_head_dim": qk_nope_head_dim, "v_head_dim": v_head_dim, "q_lora_rank": q_lora_rank,
        "num_experts": num_experts, "num_experts_per_tok": num_experts_per_tok,
        "n_dense": n_dense, "n_moe": n_moe,
        "routed_intermediate": routed_intermediate,
        "dense_intermediate": dense_intermediate,
        "n_shared_experts": n_shared, "shared_intermediate": shared_intermediate,
        "IdxD": IdxD, "idx_heads": idx_heads,
        "draft_layers": draft_layers,
        "model_type": model_type,
        "has_vision": cfg["vision"] is not None,
    }


def estimate_weights(a: dict, tp: int, ep: int, quant: str) -> dict:
    """
    W8A8/BF16 per-card weight breakdown (GiB). Family-aware:
      - MLA (DeepSeek): q_a/q_b/kv_a/kv_b/o projections via lora ranks
      - GQA (Qwen3/MiniMax/Llama): q/k/v/o via n_h*D / n_kv*D
      - MoE: routed experts /EP, shared experts /TP, gate fp32, dense MLP /TP
      - MiniMax indexer: sparse index projections /TP
    Embeddings stay BF16 (*2); linear weights use `b` (1 for W8A8, 2 for BF16).
    """
    H = a["hidden_size"]; vocab = a["vocab_size"]; L = a["L"]
    n_h = a["n_heads"]; n_kv = a["Hkv"]; D = a["D"]
    E = a["num_experts"]; n_shared = a["n_shared_experts"]
    n_dense = a["n_dense"]; n_moe = a["n_moe"]; n_sparse = a["L_sparse"]
    I_dense = a["dense_intermediate"]; I_routed = a["routed_intermediate"]
    I_shared = a["shared_intermediate"]; idx_heads = a["idx_heads"]; idx_dim = a["IdxD"]

    emb = vocab * H; lm = 0 if a["tie"] else vocab * H

    # ---- attention per layer ----
    if a["is_mla"]:
        kv_lora = a["kv_lora_rank"]; q_lora = a["q_lora_rank"]
        qk_nope = a["qk_nope_head_dim"]; qk_rope = a["qk_rope_head_dim"]; v_h = a["v_head_dim"]
        # MLA: q_a(H*q_lora)+q_norm + q_b(q_lora*n_h*(qk_nope+qk_rope))
        #      kv_a(H*(kv_lora+qk_rope))+kv_norm + kv_b(kv_lora*n_h*v_h) + o(n_h*v_h*H)
        attn = (H * q_lora + q_lora * n_h * (qk_nope + qk_rope)
                + H * (kv_lora + qk_rope) + kv_lora * n_h * v_h + n_h * v_h * H)
        qk_norm = (q_lora + kv_lora + qk_rope)  # MLA norm dims (small)
    else:
        attn = H * (n_h * D) + H * (n_kv * D) * 2 + (n_h * D) * H  # q + (k+v) + o
        qk_norm = (n_h + n_kv) * D

    rms = (2 * L + 1) * H
    indexer = (H * idx_heads * idx_dim + H * idx_dim + (idx_heads + 1) * idx_dim) if idx_dim else 0
    dense_mlp = 3 * H * I_dense if I_dense else 0
    shared = n_shared * 3 * H * I_shared if (n_shared and I_shared) else 0
    routed = E * 3 * H * I_routed if (E and I_routed) else 0
    gate = (H * E + E) if E else 0

    b = 1.0 if (quant and quant.lower() in ("ascend", "w8a8")) else 2.0
    parts = {
        "K_emb_lm_tp": round((emb + lm) / tp * 2 / GiB, 3),
        "K_norms": round((rms + L * qk_norm) * 2 / GiB, 3),
        "K_gate_fp32": round(n_moe * gate * 4 / GiB, 3) if E else 0,
        "Q_attn_tp": round(L * attn / tp * b / GiB, 3),
        "Q_dense_tp": round(n_dense * dense_mlp / tp * b / GiB, 3) if n_dense else 0,
        "Q_shared_tp": round(n_moe * shared / tp * b / GiB, 3) if shared else 0,
        "Q_indexer_tp": round(n_sparse * indexer / tp * b / GiB, 3) if idx_dim else 0,
        "Q_experts_ep": round(n_moe * routed / ep * b / GiB, 3) if E else 0,
    }
    parts["weight"] = round(sum(v for v in parts.values() if v), 3)
    return parts


# ---------- warmup log parsing (optional, for measured values) ----------
def parse_warmup_log(path: Path | None) -> dict:
    """Extract measured memory from vllm-ascend warmup log. Returns {} if no log."""
    m = {}
    if not path or not path.exists():
        return m
    text = path.read_text(encoding="utf-8", errors="ignore")
    # Loading model weights took X GB
    if x := re.search(r"Loading model weights took ([\d.]+)", text):
        m["weights_gib"] = float(x.group(1))
    # Available KV cache memory: X GiB
    if x := re.search(r"Available KV cache memory: ([\d.]+)", text):
        m["available_kv_gib"] = float(x.group(1))
    # Actual usage: X for weights, Y for peak activation, Z for non-torch, W for NPU graph
    if x := re.search(r"Actual usage: ([\d.]+) GiB for weights.*?([\d.]+) GiB for peak activation.*?([\d.]+) GiB for non-torch.*?([\d.]+) GiB for NPU graph", text):
        m["weights_gib"] = float(x.group(1)); m["peak_act_gib"] = float(x.group(2))
        m["non_torch_gib"] = float(x.group(3)); m["npu_graph_gib"] = float(x.group(4))
    # GPU KV cache size: X tokens
    if x := re.search(r"GPU KV cache size: ([\d,]+) tokens", text):
        m["gpu_kv_tokens"] = int(x.group(1).replace(",", ""))
    # Maximum concurrency for X tokens per request: Yx
    if x := re.search(r"Maximum concurrency for [\d,]+ tokens per request: ([\d.]+)x", text):
        m["max_concurrency"] = float(x.group(1))
    return m


# ---------- xlsx filling ----------
def sc(ws, r, c, v, fill=None):
    cell = ws.cell(r, c, v)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    if fill is not None:
        cell.fill = fill
    return cell


def build_row_formulas(a: dict, deploy: dict, wbk: dict, layout: KVLayout,
                       budget: MemoryBudget, hybrid: bool, num_blocks: int,
                       max_conc: float, gpu_tokens: int, blocks_per_req: int,
                       measured: dict, B: int, w_gib: float, act_gib: float,
                       nt_gib: float, g_gib: float, cur_kv: float, fit_req: float,
                       full_free: float, kv_mib: float, idx_mib: float,
                       sum_mib: float, Hkv_local: int) -> dict:
    """
    Generate per-model C-column formula text (aligned to vllm / vllm-ascend source).
    Returns {row: formula_text}. Formulas are NOT fixed — they adapt to the detected
    architecture (GQA vs MLA, MoE vs dense, shared/index/draft presence, quant scheme).
    """
    TP, EP = deploy["TP"], deploy["EP"]
    bpe = layout.bytes_per_elem
    L = a["L"]
    req_gib = round(budget.requested_bytes() / GiB, 3)
    has_log = bool(measured)
    draft_mib = round(layout.draft_bytes_per_token / MiB, 6)
    f: dict[int, str] = {}

    # Row 9 — Hkv
    if a["is_mla"]:
        f[9] = (f"MLA 不使用 Hkv（KV cache 由 kv_lora_rank 决定）；Hkv={a['Hkv']} 仅参考 "
                f"[MLAAttentionSpec; vllm_ascend MLA]")
    else:
        f[9] = (f"全局 num_key_value_heads={a['Hkv']}；TP 后 Hkv_local=max(1,Hkv//TP)"
                f"=max(1,{a['Hkv']}//{TP})={Hkv_local} [AttentionSpec TP 切分]")

    # Row 16 — experts_per_device
    if a["num_experts"]:
        f[16] = (f"num_experts // EP = {a['num_experts']}//{EP}={a['num_experts']//EP} "
                 f"[MoE EP 分片; vllm/distributed]")
    else:
        f[16] = "无 MoE → 0 [dense 模型]"

    # Row 21 — EP
    if deploy["enable_expert_parallel"]:
        f[21] = f"EP = TP×DP = {TP}×{deploy['DP']}={EP}（开 --enable-expert-parallel）[vllm-Ascend MoE EP]"
    else:
        f[21] = f"EP = TP = {TP}（未开 --enable-expert-parallel，专家按 TP 切）[vllm/distributed]"

    # Row 24 — 权重 (GiB)
    if has_log:
        f[24] = (f"实测 model_runner.model_memory_usage = {w_gib} GiB [worker.py::load_model; "
                 f"日志 'Loading model weights took ...']")
    else:
        f[24] = (f"理论估算 = {w_gib} GiB（绿底，见行39 分项）；拉起后用 warmup 日志 'Loading model "
                 f"weights took X GB' 覆盖 [worker.py::load_model]")

    # Row 29 — Current KV
    f[29] = (f"available_kv = requested − (W + peak_act + non_torch) = {req_gib} − "
             f"({w_gib}+{act_gib}+{nt_gib}) = {cur_kv} GiB ※不含 graph [worker.py:581]")

    # Row 30 — fit_requested
    f[30] = (f"requested − (W+act+non_torch+graph) − 150MiB = {req_gib} − "
             f"({w_gib}+{act_gib}+{nt_gib}+{g_gib}) − 150MiB = {fit_req} GiB [worker.py:728]")

    # Row 31 — full_free
    f[31] = (f"init_free − (W+act+non_torch+graph) − 150MiB = {budget.hbm_gib} − "
             f"({w_gib}+{act_gib}+{nt_gib}+{g_gib}) − 150MiB = {full_free} GiB [worker.py:729]")

    # Row 32 — num_blocks
    eff = layout.effective_main_layers
    if hybrid:
        gs = max(eff, a["L_sparse"] if a["L_sparse"] else eff)
        if a["is_mla"]:
            f[32] = (f"hybrid: page=block·(kv_lora+qk_rope)·dtype (MLA); "
                     f"group_size=max(L+draft,L_sparse)=max({eff},{a['L_sparse']})={gs}; "
                     f"num_blocks=avail//page//gs={num_blocks} [kv_cache_utils.py:1380]")
        else:
            f[32] = (f"hybrid: page=2·block·Hkv_local·D·dtype; "
                     f"group_size=max(L+draft,L_sparse)=max({eff},{a['L_sparse']})={gs}; "
                     f"num_blocks=avail//page//gs={num_blocks} [kv_cache_utils.py:1380]")
    else:
        if a["is_mla"]:
            f[32] = (f"uniform: page=block·(kv_lora+qk_rope)·dtype (MLA, no ×2); "
                     f"num_blocks=avail//page//(L+draft)={num_blocks} "
                     f"[kv_cache_utils.py:1009 get_num_blocks]")
        else:
            f[32] = (f"uniform: page=2·block·Hkv_local·D·dtype; "
                     f"num_blocks=avail//page//(L+draft)={num_blocks} "
                     f"[kv_cache_utils.py:1009 get_num_blocks]")

    # Row 35 — max_concurrency
    f[35] = (f"max_concurrency = num_blocks / blocks_per_req = {num_blocks}/{blocks_per_req}"
             f"={round(max_conc,4)} [kv_cache_utils.py:958]")

    # Row 39 — 权重 breakdown (per-model)
    terms = []
    if a["is_mla"]:
        attn_desc = (f"Q_attn/TP (MLA: H·q_lora + q_lora·n_h·(qk_nope+qk_rope) + "
                     f"H·(kv_lora+qk_rope) + kv_lora·n_h·v_head + n_h·v_head·H)")
    else:
        attn_desc = f"Q_attn/TP (GQA: H·n_h·D + 2·H·n_kv·D + n_h·D·H)"
    if wbk.get("K_emb_lm_tp"):
        terms.append(f"K_emb_lm/TP·2={wbk['K_emb_lm_tp']}")
    if wbk.get("K_norms"):
        terms.append(f"K_norms·2={wbk['K_norms']}")
    if wbk.get("K_gate_fp32"):
        terms.append(f"K_gate·4(fp32)={wbk['K_gate_fp32']}")
    if wbk.get("Q_attn_tp"):
        terms.append(f"{attn_desc}={wbk['Q_attn_tp']}")
    if wbk.get("Q_dense_tp"):
        terms.append(f"Q_dense/TP={wbk['Q_dense_tp']}")
    if wbk.get("Q_shared_tp"):
        terms.append(f"Q_shared/TP={wbk['Q_shared_tp']}")
    if wbk.get("Q_indexer_tp"):
        terms.append(f"Q_indexer/TP={wbk['Q_indexer_tp']}")
    if wbk.get("Q_experts_ep"):
        terms.append(f"Q_experts/EP={wbk['Q_experts_ep']}")
    f[39] = " + ".join(terms) + f" = {wbk['weight']} GiB [worker.py + 推导; 行24 同值]"

    # Row 41 — 可用KV
    f[41] = f"= 行29 = {cur_kv} GiB [worker.py:581]"

    # Row 42 — 单token kv
    if a["is_mla"]:
        f[42] = (f"L×(kv_lora_rank+qk_rope_head_dim)×dtype = {L}×({a['kv_lora_rank']}+"
                 f"{a['qk_rope_head_dim']})×{bpe} = {kv_mib} MiB (MLA, 无 ×2) "
                 f"[MLAAttentionSpec.real_page_size_bytes/block_size]")
    else:
        f[42] = (f"L×2×Hkv_local×D×dtype = {L}×2×{Hkv_local}×{a['D']}×{bpe} = {kv_mib} MiB "
                 f"[AttentionSpec.real_page_size_bytes/block_size]")

    # Row 43 — 单token index
    if a["IdxD"]:
        f[43] = (f"L_sparse×1×IdxD×dtype = {a['L_sparse']}×1×{a['IdxD']}×{bpe} = {idx_mib} MiB "
                 f"(key-only, 无 ×2) [MLAAttentionSpec; indexer.py:161]")
    else:
        f[43] = "无 sparse indexer → 0 [config 无 sparse_attention_config 或 sparse_index_dim=0]"

    # Row 44 — draft
    if a["draft_layers"]:
        per = round(layout.per_layer_main_bytes / MiB, 6)
        f[44] = (f"draft_layers×per_layer_main_kv = {a['draft_layers']}×{per} = {draft_mib} MiB "
                 f"(MTP/NextN draft 复用主层 KV) [vllm/v1/spec_decode]")
    else:
        f[44] = "无 MTP/NextN draft 层 → 0 [config 无 num_mtp_modules/num_nextn_predict_layers]"

    # Row 45 — 合计
    f[45] = (f"main + index + draft = {kv_mib}+{idx_mib}+{draft_mib} = {sum_mib} MiB "
             f"(每卡口径，含 TP 分片) [推导]")

    # Row 46 — 可容纳tokens
    f[46] = (f"max_concurrency × max_model_len = {round(max_conc,4)}×{deploy['max_model_len']} "
             f"= {gpu_tokens} [kv_cache_utils.py:1833 get_kv_cache_capacity]")

    # Row 47 — 最大并发
    if sum_mib:
        f[47] = (f"(available_kv / sum_bytes_per_token) / B = ({cur_kv} GiB / {sum_mib} MiB) / "
                 f"{B} = {round(cur_kv * GiB / (layout.sum_bytes_per_token * B), 3)} "
                 f"[kv_cache_utils.py:958]")
    else:
        f[47] = "sum_bytes_per_token=0 → 0"
    return f


def fill_xlsx(template_path: Path, out_path: Path, model_name: str, a: dict,
              deploy: dict, wbk: dict, budget: MemoryBudget, layout: KVLayout,
              hybrid: bool, num_blocks: int, max_conc: float, gpu_tokens: int,
              blocks_per_req: int, idx_bs: int, measured: dict, B: int) -> Path:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    ws.title = "Sheet1"
    for s in list(wb.sheetnames):
        if s != "Sheet1":
            del wb[s]

    Hkv_local = layout.Hkv_local
    kv_mib = round(layout.main_bytes_per_token / MiB, 6)
    idx_mib = round(layout.index_bytes_per_token / MiB, 6)
    sum_mib = round(layout.sum_bytes_per_token / MiB, 6)
    cur_kv = round(budget.current_kv_gib(), 3)
    fit_req = round(budget.kv_fit_requested_bytes() / GiB, 3)
    full_free = round(budget.kv_fully_utilize_free_bytes() / GiB, 3)
    w_gib = measured.get("weights_gib", wbk["weight"])
    act_gib = round(measured.get("peak_act_gib", budget.peak_activation_gib), 3)
    nt_gib = measured.get("non_torch_gib", budget.non_torch_gib)
    g_gib = measured.get("npu_graph_gib", budget.npu_graph_gib)
    has_log = bool(measured)
    wfill = MEAS if has_log else EST

    # 模板 A 列标签固定；B 列写数值；C 列公式由 build_row_formulas 按本模型架构动态生成
    # （覆盖模板里的通用占位），不再使用固定公式文本。
    def put(r, v, fill=None):
        cell = ws.cell(r, 2, v)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if fill is not None:
            cell.fill = fill
        return cell

    def set_c(r, text):
        cell = ws.cell(r, 3, text)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        return cell

    put(2, B, YELLOW)
    put(3, deploy["max_model_len"], YELLOW)
    put(4, deploy["util"], YELLOW)
    put(6, model_name, HEADER)
    put(7, a["L"]); put(8, a["L_sparse"]); put(9, a["Hkv"]); put(10, a["D"])
    put(11, a["IdxD"]); put(12, layout.bytes_per_elem); put(13, a["hidden_size"])
    put(14, a["num_experts_per_tok"]); put(15, a["num_experts"])
    put(16, a["num_experts"] // deploy["EP"])
    put(19, deploy["TP"], YELLOW); put(20, deploy["DP"], YELLOW); put(21, deploy["EP"], YELLOW)
    put(24, w_gib, wfill); put(25, act_gib, wfill); put(26, round(act_gib * GiB / 1e6, 1), wfill)
    put(27, nt_gib, wfill); put(28, g_gib, wfill); put(29, cur_kv, wfill)
    put(30, fit_req, wfill); put(31, full_free, wfill); put(32, num_blocks, wfill)
    put(33, layout.block_size); put(34, gpu_tokens, wfill); put(35, round(max_conc, 4), wfill)
    put(39, w_gib); put(40, act_gib); put(41, cur_kv); put(42, kv_mib)
    put(43, idx_mib); put(44, round(layout.draft_bytes_per_token / MiB, 6)); put(45, sum_mib); put(46, gpu_tokens)
    put(47, round(cur_kv * GiB / (layout.sum_bytes_per_token * B), 3) if layout.sum_bytes_per_token else 0)

    # C 列：按本模型架构动态生成公式（覆盖模板通用占位）
    formulas = build_row_formulas(a, deploy, wbk, layout, budget, hybrid, num_blocks,
                                  max_conc, gpu_tokens, blocks_per_req, measured, B,
                                  w_gib, act_gib, nt_gib, g_gib, cur_kv, fit_req, full_free,
                                  kv_mib, idx_mib, sum_mib, Hkv_local)
    for r, text in formulas.items():
        set_c(r, text)

    asum = wb.create_sheet("假设与说明")
    rows = [("项目", "说明"), ("HBM", f"{budget.hbm_gib} GiB"), ("量化", deploy.get("quantization") or "BF16"),
            ("权重/激活/non_torch/graph", "实测(蓝底)" if has_log else "理论占位(绿底)，需warmup日志替换"),
            ("关键结果", f"W={w_gib}GiB, CurrentKV={cur_kv}GiB, tokens={gpu_tokens}, 满长并发={round(max_conc,4)}"),
            ("源码版本", "vllm d6dbdb9b0 / vllm-ascend f5b5514af")]
    for i, (x, y) in enumerate(rows, 1):
        asum.cell(i, 1, x); asum.cell(i, 2, y)
        if i == 1: asum.cell(i, 1).fill = HEADER; asum.cell(i, 2).fill = HEADER
    asum.column_dimensions["A"].width = 28; asum.column_dimensions["B"].width = 90
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(
        description="Ascend NPU 显存填表（无需本地 vllm/vllm-ascend 源码；公式已内置）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
示例:
  # 1) 只给 HF model id（自动下载 config.json，无需本地任何模型文件）
  python fill_table.py --hf-model MiniMaxAI/MiniMax-M3 \\
    --serve "vllm serve MiniMaxAI/MiniMax-M3 --tensor-parallel-size 8 --gpu-memory-utilization 0.9 --max-model-len 65536 --quantization ascend --enable-expert-parallel"

  # 2) 从 ModelScope 下载（国内更快；或 HF 不可达时自动回退到 ModelScope）
  python fill_table.py --ms-model Qwen/Qwen3-235B-A22B --serve "vllm serve ..."

  # 3) 本地 config.json
  python fill_table.py --config /path/to/config.json --serve "vllm serve ..." --hbm 96

  # 4) 只算不写表（dry-run，快速看结果）
  python fill_table.py --hf-model Qwen/Qwen3-235B-A22B --serve "..." --dry-run

  # 5) 带 warmup 日志（实测值覆盖理论占位）
  python fill_table.py --config config.json --serve "..." --warmup-log vllm.log

  # 6) 强制数据源：--source hf|ms|auto（默认 auto：先 HF 后 ModelScope）
  python fill_table.py --hf-model deepseek-ai/DeepSeek-V3 --source ms --serve "..."
""",
    )
    src = ap.add_argument_group("模型来源（四选一）")
    src.add_argument("--hf-model", help="HuggingFace model id 或 URL（如 MiniMaxAI/MiniMax-M3）；自动下载 config.json，无需权重")
    src.add_argument("--ms-model", help="ModelScope model id 或 URL（如 Qwen/Qwen3-235B-A22B）；国内推荐")
    src.add_argument("--config", help="本地 config.json 路径")
    src.add_argument("--config-dir", help="含 config.json 的模型目录")
    src.add_argument("--source", choices=["auto", "hf", "ms"], default="auto",
                     help="config.json 下载源：auto=先 HF 后 ModelScope（默认）；hf=仅 HF；ms=仅 ModelScope")
    ap.add_argument("--serve", required=True, help="vllm serve / api_server 拉起命令字符串")
    ap.add_argument("--model-name", default="", help="表格中显示的模型名（不给则用 model id / 目录名）")
    ap.add_argument("--hbm", type=float, default=64.0, help="单卡 HBM GiB（默认 64；可选 80/96）")
    ap.add_argument("--B", type=int, default=8192, help="平均请求长度（输入+输出 tokens，默认 8192）")
    ap.add_argument("--non-torch", type=float, default=3.2, help="non-torch 经验值 GiB（TP/DP 越大越大，默认 3.2）")
    ap.add_argument("--graph", type=float, default=2.0, help="NPU graph 经验值 GiB（默认 2.0）")
    ap.add_argument("--warmup-log", help="vllm-ascend warmup 日志路径（可选，提供则用实测值）")
    ap.add_argument("--template", help="模板 xlsx（默认用脚本同目录 template.xlsx）")
    ap.add_argument("--out", default=None, help="输出 xlsx 路径（默认写到 skill 目录下 outputs/<模型名>-memory-table.xlsx）")
    ap.add_argument("--dry-run", action="store_true", help="只计算并打印结果，不写 xlsx")
    args = ap.parse_args()

    # ---- resolve config.json source ----
    if not args.hf_model and not args.ms_model and not args.config and not args.config_dir:
        sys.exit("[fill_table] 必须提供模型来源之一：--hf-model / --ms-model / --config / --config-dir")

    tmp_dir = Path(tempfile.gettempdir()) / "ascend-mem-table"

    if args.hf_model or args.ms_model:
        raw_id = args.hf_model or args.ms_model
        model_id = _hf_resolve_id(raw_id)
        # --ms-model forces ms; --hf-model forces hf; --source overrides
        if args.ms_model:
            source = "ms"
        elif args.hf_model:
            source = "hf" if args.source == "auto" else args.source
        else:
            source = args.source
        # for --hf-model with --source auto, allow fallback to ms
        if args.hf_model and args.source == "auto":
            source = "auto"
        tmp_dir_inner = Path(tempfile.gettempdir()) / "ascend-mem-table"
        cfg_path, used = download_config(model_id, tmp_dir_inner, source=source)
        if used == "ms":
            print(f"[fill_table] config.json 来自 ModelScope（HF 不可达或 --source ms）")
        default_name = model_id.split("@", 1)[0].split("/", 1)[-1]
        prefer_source = used
        local_config_dir = None
    else:
        cfg_path = Path(args.config) if args.config else Path(args.config_dir) / "config.json"
        if not cfg_path.exists():
            sys.exit(f"[fill_table] config.json 不存在: {cfg_path}")
        default_name = cfg_path.parent.name
        model_id = None
        prefer_source = None
        local_config_dir = cfg_path.parent
        source = args.source

    cfg = load_config(cfg_path)
    a = arch_from_config(cfg)
    deploy = parse_serve(args.serve)
    if not deploy["max_model_len"]:
        # fall back to config max_position_embeddings if serve didn't specify
        mpe = cfg["text"].get("max_position_embeddings") or cfg["text"].get("seq_length")
        deploy["max_model_len"] = mpe or 8192
        print(f"[fill_table] 未在拉起命令中找到 --max-model-len，回退到 {deploy['max_model_len']}")
    if not deploy["max_num_batched_tokens"]:
        deploy["max_num_batched_tokens"] = deploy["max_model_len"]
    name = args.model_name or default_name

    quant = deploy.get("quantization") or ""
    is_w8a8 = quant.lower() in ("ascend", "w8a8")
    wbk = estimate_weights(a, deploy["TP"], deploy["EP"], quant)
    Hkv_local = local_num_kv_heads(a["Hkv"], deploy["TP"])
    n_h_local = max(1, a["n_heads"] // deploy["TP"])
    T = deploy["max_num_batched_tokens"]
    act_b = estimate_peak_activation_bytes(T, a["hidden_size"], n_h_local, a["D"], a["num_experts_per_tok"], deploy["EP"])
    # KV cache dtype auto-detection: --kv-cache-dtype > quant_model_description.json (C8) > vLLM default (BF16)
    bpe, bpe_reason = detect_bpe(deploy, model_id, tmp_dir, source, prefer_source, local_config_dir)
    print(f"[fill_table] KV cache dtype → bytes_per_elem={bpe}  ({bpe_reason})")
    budget = MemoryBudget(args.hbm, deploy["util"], wbk["weight"], act_b / GiB, args.non_torch, args.graph, args.hbm)
    layout = KVLayout(
        a["L"], a["L_sparse"], a["Hkv"], Hkv_local, a["D"], a["IdxD"], bpe,
        deploy["block_size"], deploy["TP"],
        is_mla=a["is_mla"], kv_lora_rank=a["kv_lora_rank"], qk_rope_head_dim=a["qk_rope_head_dim"],
        draft_layers=a["draft_layers"],
    )
    hybrid = a["L_sparse"] > 0
    avail = budget.available_kv_bytes()
    num_blocks = layout.num_blocks_hybrid_m3(avail) if hybrid else layout.num_blocks_uniform(avail)
    max_conc, gpu_tokens, blocks_per_req, idx_bs = layout.concurrency_and_tokens(num_blocks, deploy["max_model_len"], hybrid)
    measured = parse_warmup_log(Path(args.warmup_log) if args.warmup_log else None)
    if measured:
        if "available_kv_gib" in measured:
            budget2 = MemoryBudget(args.hbm, deploy["util"], measured.get("weights_gib", wbk["weight"]),
                                   measured.get("peak_act_gib", act_b / GiB), measured.get("non_torch_gib", args.non_torch),
                                   measured.get("npu_graph_gib", args.graph), args.hbm)
            avail = budget2.available_kv_bytes()
            num_blocks = layout.num_blocks_hybrid_m3(avail) if hybrid else layout.num_blocks_uniform(avail)
            max_conc, gpu_tokens, blocks_per_req, idx_bs = layout.concurrency_and_tokens(num_blocks, deploy["max_model_len"], hybrid)
            budget = budget2
        if "gpu_kv_tokens" in measured:
            gpu_tokens = measured["gpu_kv_tokens"]
        if "max_concurrency" in measured:
            max_conc = measured["max_concurrency"]

    cur_kv = round(budget.current_kv_gib(), 3)
    print(f"模型: {name}  (model_type={a['model_type'] or 'unknown'}"
          + (" MLA" if a["is_mla"] else "") + (" MTP" if a["draft_layers"] else "") + ")")
    print(f"  L={a['L']} L_sparse={a['L_sparse']} Hkv={a['Hkv']} D={a['D']} IdxD={a['IdxD']} "
          f"hidden={a['hidden_size']} experts={a['num_experts']} topk={a['num_experts_per_tok']} "
          f"n_dense={a['n_dense']} n_moe={a['n_moe']} n_shared={a['n_shared_experts']} draft={a['draft_layers']}"
          + (f" kv_lora={a['kv_lora_rank']} qk_rope={a['qk_rope_head_dim']}" if a["is_mla"] else ""))
    print(f"  部署: TP={deploy['TP']} DP={deploy['DP']} EP={deploy['EP']} util={deploy['util']} "
          f"max_model_len={deploy['max_model_len']} block_size={deploy['block_size']} quant={quant or 'BF16'} kv_dtype_bpe={bpe}")
    if a["is_mla"]:
        print(f"  [MLA] per-layer KV = (kv_lora_rank+qk_rope_head_dim)×dtype = ({a['kv_lora_rank']}+{a['qk_rope_head_dim']})×{bpe}")
    else:
        print(f"  Hkv_local=max(1,{a['Hkv']}//{deploy['TP']})={Hkv_local}")
    print(f"  权重分项: " + " ".join(f"{k}={v}" for k, v in wbk.items() if k != "weight"))
    print(f"  W={wbk['weight']}GiB  peak_act={act_b/GiB:.3f}GiB  non_torch={args.non_torch}GiB  graph={args.graph}GiB")
    print(f"  requested={budget.requested_bytes()/GiB:.3f}GiB  CurrentKV={cur_kv}GiB  "
          f"num_blocks={num_blocks}  blocks_per_req={blocks_per_req}")
    print(f"  GPU_KV_tokens={gpu_tokens}  满长并发={max_conc:.4f}  "
          f"单token合计={layout.sum_bytes_per_token/MiB:.6f}MiB "
          f"(main={layout.main_bytes_per_token/MiB:.6f} index={layout.index_bytes_per_token/MiB:.6f} draft={layout.draft_bytes_per_token/MiB:.6f})")
    if measured:
        print(f"  [实测] 已用 warmup 日志覆盖: {list(measured.keys())}")
    else:
        print(f"  [理论] 无 warmup 日志，权重/激活/non_torch/graph 为理论占位（绿底）；拉起后可用日志替换")

    if args.dry_run:
        print("[dry-run] 未写 xlsx")
        return

    # 默认输出到 skill 目录下的 outputs/ 子文件夹，避免污染 skill 结构
    if args.out:
        out_path = Path(args.out)
    else:
        out_path = Path(__file__).resolve().parent.parent / "outputs" / f"{name}-memory-table.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tpl = Path(args.template) if args.template else Path(__file__).parent / "template.xlsx"
    if not tpl.exists():
        sys.exit(f"[fill_table] 模板不存在: {tpl}（请先运行 python build_template.py 生成）")
    out = fill_xlsx(tpl, out_path, name, a, deploy, wbk, budget, layout, hybrid,
                    num_blocks, max_conc, gpu_tokens, blocks_per_req, idx_bs, measured, args.B)
    print(f"OK -> {out}")


if __name__ == "__main__":
    main()
